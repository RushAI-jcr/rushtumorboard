#!/usr/bin/env python3
"""
Parse any Tumor Board Excel export into per-patient CSV folders.

Usage:
    python parse_tumor_board_excel.py                          # auto-detect Excel in parent dir
    python parse_tumor_board_excel.py /path/to/TumorBoard.xlsx  # explicit path
    python parse_tumor_board_excel.py --dry-run                # preview only, no writes

Creates one folder per PatientID under infra/patient_data/{patient_id}/ with:
  - 7 standard Caboodle CSVs (clinical_notes, pathology_reports, radiology_reports,
    lab_results, cancer_staging, medications, diagnoses)
  - variant_details.csv (if VARIANT DETAILS sheet exists)
  - variant_interpretation.csv (if VARIANT INTERPRETATION sheet exists)
  - patient_demographics.csv (MRN extracted from clinical notes/reports)

Skips synthetic test patients (patient_gyn_001, patient_gyn_002, patient_4).
Existing patient folders are updated (new CSVs overwrite); demographics are only
written if patient_demographics.csv doesn't already exist.
"""

import argparse
import csv
import datetime
import glob
import os
import re
import sys
from collections import Counter

import openpyxl

# Paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(SCRIPT_DIR)
DEFAULT_EXCEL_DIR = os.path.dirname(REPO_ROOT)  # ../localtumorboard/
OUTPUT_DIR = os.path.join(REPO_ROOT, "infra", "patient_data")

# Synthetic patients to skip (don't overwrite)
SKIP_PATIENTS = {"patient_gyn_001", "patient_gyn_002", "patient_4"}

# Sheet name -> output CSV filename.
# Keys are checked case-insensitively against the workbook's sheet names.
SHEET_MAP = {
    "clinical_notes.csv": "clinical_notes.csv",
    "pathology_reports.csv": "pathology_reports.csv",
    "radiology_reports.csv": "radiology_reports.csv",
    "lab_results.csv": "lab_results.csv",
    "cancer_staging.csv": "cancer_staging.csv",
    "medications.csv": "medications.csv",
    "diagnoses.csv": "diagnoses.csv",
    "variant details": "variant_details.csv",
    "variant interpretation": "variant_interpretation.csv",
}

# MRN regex patterns (ordered from most specific to broadest)
MRN_PATTERNS = [
    re.compile(r"MRN[:\s#]*(\d{5,10})", re.IGNORECASE),
    re.compile(r"Medical Record (?:Number|No\.?)[:\s#]*(\d{5,10})", re.IGNORECASE),
    re.compile(r"Med\.?\s*Rec\.?[:\s#]*(\d{5,10})", re.IGNORECASE),
]

# 7-digit numbers that are NOT MRNs (zip codes, CPT codes, phone prefixes, etc.)
NON_MRN_NUMBERS = frozenset({
    "60612", "60616", "60607", "60608", "60637",  # Chicago zip codes
    "88300", "88302", "88304", "88305", "88307", "88309", "88321", "88323",  # CPT path
    "88342", "88360", "88361", "88364", "88365", "88367",  # CPT IHC
    "3122",  # Chicago area code prefix
})


def format_value(val):
    """Convert cell value to CSV-safe string."""
    if val is None:
        return ""
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")
    return str(val)


def find_patient_id_col(headers):
    """Find the index of the PatientID column (case-insensitive)."""
    for i, h in enumerate(headers):
        if h and "patientid" in str(h).lower().replace("_", "").replace(" ", ""):
            return i
    return None


def match_sheet_name(sheet_name):
    """Match a workbook sheet name to our SHEET_MAP (case-insensitive).

    Returns the CSV filename if matched, else None.
    """
    lower = sheet_name.strip().lower()
    for key, csv_name in SHEET_MAP.items():
        if lower == key.lower():
            return csv_name
    return None


def parse_sheet(ws):
    """Parse a worksheet into {patient_id: [rows]} grouped by PatientID.

    Returns (headers, patient_rows) where patient_rows is a dict.
    """
    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))

    pat_col = find_patient_id_col(headers)
    if pat_col is None:
        print(f"  WARNING: No PatientID column found, skipping")
        return headers, {}

    patient_rows = {}
    for row in rows_iter:
        patient_id = row[pat_col]
        if patient_id is None or str(patient_id).strip() == "":
            continue
        pid = str(patient_id).strip()
        if pid not in patient_rows:
            patient_rows[pid] = []
        patient_rows[pid].append(row)

    return headers, patient_rows


def extract_mrn_from_notes(sheet_data):
    """Extract MRN from clinical notes, pathology, and radiology text.

    First tries explicit MRN patterns. For patients with no match, falls back
    to the most frequent 7-digit number in their notes (filtering out known
    non-MRN numbers like zip codes and CPT codes).
    """
    mrn_counts = {}  # pid -> Counter of MRN candidates (explicit patterns)

    # Sheets with free-text columns to scan
    text_columns = {
        "clinical_notes.csv": ("CONCATENATED_TEXT", "NoteText", "TEXT"),
        "pathology_reports.csv": ("ReportText", "CONCATENATED_TEXT", "TEXT"),
        "radiology_reports.csv": ("ReportText", "CONCATENATED_TEXT", "TEXT"),
    }

    # Collect all text per patient for fallback
    patient_text = {}  # pid -> list of text strings

    for csv_name, text_headers in text_columns.items():
        # Find which sheet_data key maps to this CSV name
        matching_key = None
        for key, (headers, patient_rows) in sheet_data.items():
            mapped = match_sheet_name(key)
            if mapped == csv_name:
                matching_key = key
                break
        if matching_key is None:
            continue

        headers, patient_rows = sheet_data[matching_key]

        # Find text column (try multiple possible names)
        text_col = None
        for candidate in text_headers:
            for i, h in enumerate(headers):
                if h and str(h).strip().upper() == candidate.upper():
                    text_col = i
                    break
            if text_col is not None:
                break
        if text_col is None:
            continue

        for pid, rows in patient_rows.items():
            if pid not in mrn_counts:
                mrn_counts[pid] = Counter()
            if pid not in patient_text:
                patient_text[pid] = []
            for row in rows:
                text = str(row[text_col]) if row[text_col] else ""
                patient_text[pid].append(text)
                for pattern in MRN_PATTERNS:
                    for m in pattern.findall(text):
                        mrn_counts[pid][m] += 1

    # Resolve: pick most frequent MRN per patient (explicit patterns)
    resolved = {}
    for pid, counter in mrn_counts.items():
        if counter:
            resolved[pid] = counter.most_common(1)[0][0]

    # Fallback: for patients with no explicit MRN, find most frequent 7-digit number
    all_pids = set(mrn_counts.keys()) | set(patient_text.keys())
    digit_re = re.compile(r"\b(\d{7})\b")
    for pid in all_pids:
        if pid in resolved:
            continue
        texts = patient_text.get(pid, [])
        if not texts:
            continue
        digit_counts = Counter()
        for text in texts:
            for m in digit_re.findall(text):
                # Filter known non-MRN numbers
                if m not in NON_MRN_NUMBERS and not any(m.startswith(p) for p in ("312", "773", "630", "847")):
                    digit_counts[m] += 1
        if digit_counts:
            best, count = digit_counts.most_common(1)[0]
            if count >= 2:  # require at least 2 occurrences for confidence
                resolved[pid] = best

    return resolved


def extract_demographics_from_notes(sheet_data):
    """Extract PatientName, DOB, and Sex from clinical note text.

    Returns {pid: {"name": ..., "dob": ..., "sex": ...}}.
    """
    name_patterns = [
        # "LAST, FIRST" in note headers (Epic format) — stop at whitespace run or newline
        re.compile(r"PATIENT\s*NAME[:\s]+([A-Z][A-Z'-]+,\s*[A-Z][A-Za-z'-]+(?:\s+[A-Z][A-Za-z'-]+)?)\s{2,}", re.IGNORECASE),
        # "Patient Name: First Last\n" (discharge/op notes) — stop at newline
        re.compile(r"Patient\s+Name:\s*([A-Z][a-z]+(?:\s+[A-Z]\.?)?\s+[A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)\s*\n", re.IGNORECASE),
    ]
    dob_patterns = [
        re.compile(r"(?:DOB|Date of Birth|D\.O\.B\.)[:\s]+(\d{1,2}/\d{1,2}/(?:19|20)\d{2})", re.IGNORECASE),
        re.compile(r"(?:DOB|Date of Birth)[:\s]+(\d{4}-\d{2}-\d{2})", re.IGNORECASE),
    ]
    sex_pattern = re.compile(r"\b(\d{2,3})\s*(?:year|yr|y\.?o\.?)\s*old\s+(fe)?male\b", re.IGNORECASE)

    # Find the clinical notes sheet
    notes_key = None
    for key, (headers, patient_rows) in sheet_data.items():
        if match_sheet_name(key) == "clinical_notes.csv":
            notes_key = key
            break
    if notes_key is None:
        return {}

    headers, patient_rows = sheet_data[notes_key]
    text_col = None
    for i, h in enumerate(headers):
        if h and str(h).strip().upper() in ("CONCATENATED_TEXT", "NOTETEXT", "TEXT"):
            text_col = i
            break
    if text_col is None:
        return {}

    results = {}
    for pid, rows in patient_rows.items():
        info = {"name": "", "dob": "", "sex": ""}
        for row in rows[:20]:  # check first 20 notes
            text = str(row[text_col]) if row[text_col] else ""
            text_head = text[:5000]

            if not info["name"]:
                for pat in name_patterns:
                    m = pat.search(text_head)
                    if m:
                        candidate = m.group(1).strip().rstrip(",").strip()
                        if 3 < len(candidate) < 50 and not candidate.startswith("{"):
                            info["name"] = candidate
                            break

            if not info["dob"]:
                for pat in dob_patterns:
                    m = pat.search(text_head)
                    if m:
                        info["dob"] = m.group(1).strip()
                        break

            if not info["sex"]:
                m = sex_pattern.search(text_head)
                if m:
                    info["sex"] = "Female" if m.group(2) else "Male"

            if info["name"] and info["dob"] and info["sex"]:
                break

        results[pid] = info
    return results


def validate_mrn_consistency(mrn_map, output_dir, mrn_map_file=None):
    """Cross-validate extracted MRNs against existing demographics and optional mapping file.

    Prints warnings for mismatches. Returns nothing.
    """
    mismatches = []

    # Check against existing patient_demographics.csv files
    for pid, extracted_mrn in mrn_map.items():
        demo_path = os.path.join(output_dir, pid, "patient_demographics.csv")
        if not os.path.isfile(demo_path):
            continue
        with open(demo_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            row = next(reader, {})
        existing_mrn = row.get("MRN", "").strip()
        if existing_mrn and extracted_mrn and existing_mrn != extracted_mrn:
            mismatches.append((pid, existing_mrn, extracted_mrn))

    # Check against optional MRN mapping file
    if mrn_map_file and os.path.isfile(mrn_map_file):
        with open(mrn_map_file, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                pid = row.get("PatientID", "").strip()
                expected_mrn = row.get("MRN", "").strip()
                extracted_mrn = mrn_map.get(pid, "")
                if expected_mrn and extracted_mrn and expected_mrn != extracted_mrn:
                    mismatches.append((pid, expected_mrn, extracted_mrn))

    if mismatches:
        print(f"\nWARNING: {len(mismatches)} MRN mismatch(es) detected:")
        for pid, expected, got in mismatches:
            exp_masked = f"***{expected[-4:]}" if len(expected) >= 4 else expected
            got_masked = f"***{got[-4:]}" if len(got) >= 4 else got
            print(f"  {pid}: existing/expected={exp_masked}, extracted={got_masked}")
        print("  -> Verify correct MRN before running agents")


def print_data_gap_report(new_patients, output_dir):
    """Print a per-patient summary of available vs missing data."""
    all_csv_types = [
        "clinical_notes", "pathology_reports", "radiology_reports",
        "lab_results", "cancer_staging", "medications", "diagnoses",
    ]

    print(f"\n{'='*80}")
    print("DATA GAP REPORT")
    print(f"{'='*80}")
    print(f"{'Patient':<40s} {'notes':>6s} {'path':>5s} {'rad':>5s} {'labs':>5s} {'stage':>6s} {'meds':>5s} {'dx':>5s}")
    print("-" * 80)

    for pid in new_patients:
        patient_dir = os.path.join(output_dir, pid)
        counts = {}
        for ft in all_csv_types:
            csv_path = os.path.join(patient_dir, f"{ft}.csv")
            if os.path.isfile(csv_path):
                with open(csv_path, "r", encoding="utf-8-sig") as f:
                    counts[ft] = sum(1 for _ in csv.DictReader(f))
            else:
                counts[ft] = -1  # missing

        def fmt(n):
            return "--" if n < 0 else str(n)

        print(
            f"{pid[:38]:<40s} "
            f"{fmt(counts['clinical_notes']):>6s} "
            f"{fmt(counts['pathology_reports']):>5s} "
            f"{fmt(counts['radiology_reports']):>5s} "
            f"{fmt(counts['lab_results']):>5s} "
            f"{fmt(counts['cancer_staging']):>6s} "
            f"{fmt(counts['medications']):>5s} "
            f"{fmt(counts['diagnoses']):>5s}"
        )

    print("-" * 80)
    print("'--' = CSV file missing (agent uses 3-layer fallback to clinical notes)")


def write_csv(filepath, headers, rows):
    """Write rows to a CSV file."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([format_value(v) for v in row])


def write_demographics_csv(filepath, patient_id, mrn, name="", dob="", sex=""):
    """Write patient_demographics.csv matching existing schema."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["PatientID", "MRN", "PatientName", "DOB", "Sex"])
        writer.writerow([patient_id, mrn, name, dob, sex])


def auto_detect_excel(directory):
    """Find the most recent Tumor Board Excel file in the given directory."""
    patterns = [
        os.path.join(directory, "Tumor Board Data*.xlsx"),
        os.path.join(directory, "tumor_board*.xlsx"),
        os.path.join(directory, "TB_*.xlsx"),
    ]
    candidates = []
    for pattern in patterns:
        candidates.extend(glob.glob(pattern))
    if not candidates:
        return None
    # Return most recently modified
    return max(candidates, key=os.path.getmtime)


def main():
    parser = argparse.ArgumentParser(
        description="Parse Tumor Board Excel export into per-patient CSV folders."
    )
    parser.add_argument(
        "excel_path", nargs="?", default=None,
        help="Path to the Tumor Board Excel file. Auto-detects if omitted."
    )
    parser.add_argument(
        "--output-dir", default=OUTPUT_DIR,
        help=f"Output directory for patient folders (default: {OUTPUT_DIR})"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Preview what would be created without writing files."
    )
    parser.add_argument(
        "--force-demographics", action="store_true",
        help="Overwrite existing patient_demographics.csv (re-extract MRN/Name/DOB)."
    )
    parser.add_argument(
        "--mrn-map",
        help="Optional CSV file with PatientID,MRN columns for MRN cross-validation."
    )
    args = parser.parse_args()

    # Resolve Excel path
    excel_path = args.excel_path
    if excel_path is None:
        excel_path = auto_detect_excel(DEFAULT_EXCEL_DIR)
        if excel_path is None:
            print("ERROR: No Tumor Board Excel file found. Provide a path as argument.")
            sys.exit(1)
        print(f"Auto-detected: {excel_path}")
    elif not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found: {excel_path}")
        sys.exit(1)

    output_dir = args.output_dir

    print(f"Reading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    print(f"Sheets: {wb.sheetnames}")

    all_patients = set()
    sheet_data = {}  # {original_sheet_name: (headers, {pid: [rows]})}

    for sheet_name in wb.sheetnames:
        csv_name = match_sheet_name(sheet_name)
        if csv_name is None:
            print(f"  Skipping unknown sheet: {sheet_name}")
            continue

        print(f"\nParsing sheet: {sheet_name} -> {csv_name}")
        ws = wb[sheet_name]
        headers, patient_rows = parse_sheet(ws)
        sheet_data[sheet_name] = (headers, patient_rows)

        for pid, rows in patient_rows.items():
            all_patients.add(pid)
            print(f"  {pid}: {len(rows)} rows")

    wb.close()

    # Extract MRN
    print(f"\n{'='*60}")
    print("Extracting MRN from clinical notes/reports...")
    mrn_map = extract_mrn_from_notes(sheet_data)

    print(f"\nPatientID -> MRN mapping:")
    for pid in sorted(all_patients):
        mrn = mrn_map.get(pid, "[NOT FOUND]")
        mrn_masked = f"***{mrn[-4:]}" if len(mrn) >= 4 else mrn
        print(f"  {pid} -> {mrn_masked}")

    # Cross-validate MRNs
    validate_mrn_consistency(mrn_map, output_dir, args.mrn_map)

    # Extract demographics (name, DOB, sex) from notes
    print(f"\nExtracting demographics from clinical notes...")
    demo_map = extract_demographics_from_notes(sheet_data)
    for pid in sorted(all_patients):
        info = demo_map.get(pid, {})
        name = info.get("name", "")
        dob = info.get("dob", "")
        sex = info.get("sex", "")
        if name or dob or sex:
            name_masked = f"{name[0]}." if name else "?"
            print(f"  {pid} -> Name={name_masked}, DOB={'***' if dob else '?'}, Sex={sex or '?'}")

    # Determine new vs existing
    new_patients = sorted(all_patients - SKIP_PATIENTS)
    existing = set()
    for pid in new_patients:
        if os.path.exists(os.path.join(output_dir, pid)):
            existing.add(pid)

    print(f"\n{'='*60}")
    print(f"Total unique patients: {len(all_patients)}")
    print(f"New patient folders to create: {len(set(new_patients) - existing)}")
    print(f"Existing folders to update: {len(existing)}")
    if args.dry_run:
        print("DRY RUN — no files will be written")
    print(f"{'='*60}")

    if args.dry_run:
        print("\nDry run complete.")
        return

    # Write CSVs
    for pid in new_patients:
        patient_dir = os.path.join(output_dir, pid)
        action = "Updating" if pid in existing else "Creating"
        print(f"\n{action}: {patient_dir}/")

        for sheet_name in sheet_data:
            csv_name = match_sheet_name(sheet_name)
            if csv_name is None:
                continue
            headers, patient_rows = sheet_data[sheet_name]
            rows = patient_rows.get(pid, [])
            if not rows:
                continue
            csv_path = os.path.join(patient_dir, csv_name)
            write_csv(csv_path, headers, rows)
            print(f"  {csv_name}: {len(rows)} rows")

        # Write demographics (only if doesn't already exist, or --force-demographics)
        mrn = mrn_map.get(pid, "")
        demo_info = demo_map.get(pid, {})
        demo_path = os.path.join(patient_dir, "patient_demographics.csv")
        if not os.path.exists(demo_path) or args.force_demographics:
            action_word = "OVERWRITING" if os.path.exists(demo_path) else "Creating"
            write_demographics_csv(
                demo_path, pid, mrn,
                name=demo_info.get("name", ""),
                dob=demo_info.get("dob", ""),
                sex=demo_info.get("sex", ""),
            )
            mrn_display = f"***{mrn[-4:]}" if len(mrn) >= 4 else (mrn or "[NOT FOUND - needs manual entry]")
            name_raw = demo_info.get("name", "")
            name_display = f"{name_raw[0]}." if name_raw else "?"
            print(f"  patient_demographics.csv: {action_word} MRN={mrn_display}, Name={name_display}")
        else:
            print(f"  patient_demographics.csv: EXISTS (skipped, use --force-demographics to overwrite)")

    # Summary
    missing = [pid for pid in new_patients if pid not in mrn_map]
    print(f"\nDone! Processed {len(new_patients)} patients in {output_dir}")
    if missing:
        print(f"\nWARNING: {len(missing)} patients have no MRN in notes:")
        for pid in missing:
            print(f"  {pid}")
        print("  -> Add MRN manually to patient_demographics.csv")

    # Data gap report
    print_data_gap_report(new_patients, output_dir)

    # Update local_patient_ids.json
    ids_path = os.path.join(REPO_ROOT, "src", "tests", "local_patient_ids.json")
    if os.path.exists(ids_path):
        import json
        with open(ids_path) as f:
            existing_ids = set(json.load(f))
        new_ids = existing_ids | set(new_patients)
        if new_ids != existing_ids:
            with open(ids_path, "w") as f:
                json.dump(sorted(new_ids), f, indent=2)
                f.write("\n")
            print(f"\nUpdated local_patient_ids.json: {len(existing_ids)} -> {len(new_ids)} patients")


if __name__ == "__main__":
    main()
